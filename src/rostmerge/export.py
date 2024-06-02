from collections.abc import Sequence
from typing import Any

from openpyxl import Workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.formatting.rule import Rule
from openpyxl.styles import PatternFill
from openpyxl.styles.differential import DifferentialStyle

HEADER = ('Last', 'First', 'usatf Age', 'USATF Status', 'USATF Num', 'Age Verified',
            'Last Mismatch', 'First Mismatch', 'DOB Mismatch', 'Gender Mismatch')
GREEN_FILL = PatternFill(start_color='B7E1CD', end_color='B7E1CD', fill_type='solid')
RED_FILL = PatternFill(start_color='E06666', end_color='E06666', fill_type='solid')


def export_roster(filename:str, roster: Sequence[Any], age_year: int) -> None:

    wb = Workbook()
    wb.active.title = 'registration'
    ws = wb.active
    for c, h in zip('ABCDEFGHIJ', HEADER, strict=True):
        ws[f'{c}1'] = CellRichText(TextBlock(InlineFont(b=True), h))
        ws.column_dimensions[c].width = 14.29
    ws.freeze_panes = 'A2'
    for row in roster:
        r = list(row[:2])
        r.append(age_year - row[2].year)
        r.append('Current' if row[3] else 'Not Assoc')
        r.append(row[4])
        r.append('Current' if row[5] else '')
        r.append(str(not row[6]))
        r.append(str(not row[7]))
        r.append(str(not row[8]))
        r.append(str(not row[9]))
        ws.append(r)
    for col in 'DF':
        rng = f'{col}2:{col}{len(roster)+1}'
        ws.conditional_formatting.add(
            rng,
            Rule(
                type='containsText', operator='containsText', text='Current',
                formula=[f'NOT(ISERROR(SEARCH("Current",{col}2)))'],
                stopIfTrue=False,
                dxf=DifferentialStyle(fill=GREEN_FILL)
            )
        )
        ws.conditional_formatting.add(
            rng,
            Rule(
                type='notContainsText', operator='notContains', text='Current',
                formula=[f'ISERROR(SEARCH("Current",{col}2))'],
                stopIfTrue=False,
                dxf=DifferentialStyle(fill=RED_FILL)
            )
        )
    for col in 'GHIJ':
        rng = f'{col}2:{col}{len(roster)+1}'
        ws.conditional_formatting.add(
            rng,
            Rule(
                type='containsText', operator='containsText', text='False',
                formula=[f'NOT(ISERROR(SEARCH("False",{col}2)))'],
                stopIfTrue=False,
                dxf=DifferentialStyle(fill=GREEN_FILL)
            )
        )
        ws.conditional_formatting.add(
            rng,
            Rule(
                type='notContainsText', operator='notContains', text='False',
                formula=[f'ISERROR(SEARCH("FALSE",{col}2))'],
                stopIfTrue=False,
                dxf=DifferentialStyle(fill=RED_FILL)
            )
        )
    wb.save(filename)
    wb.close()
