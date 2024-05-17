from contextlib import closing
import csv
from datetime import date, datetime
import sqlite3
import sys

from openpyxl import Workbook
from openpyxl.cell.text import InlineFont
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.formatting.rule import Rule
from openpyxl.styles import PatternFill
from openpyxl.styles.differential import DifferentialStyle


ROSTER_SCHEMA = """
CREATE TABLE IF NOT EXISTS roster (
    ts_last TEXT,
    ts_first TEXT,
    ts_dob DATE,
    ts_gender TEXT,
    usatf_last TEXT,
    usatf_first TEST,
    usatf_dob DATE,
    usatf_id INTEGER,
    usatf_valid BOOLEAN DEFAULT FALSE,
    usatf_age_verified BOOLEAN DEFAULT FALSE,
    is_coach BOOLEAN DEFAULT FALSE,
    touched TIMESTAMP,
    UNIQUE(ts_last, ts_first, ts_dob)
);
"""


USATF_SCHEMA = """
CREATE TEMPORARY TABLE usatf (
    usatf_last TEXT,
    usatf_first TEXT,
    usatf_dob DATE,
    usatf_gender TEXT,
    usatf_id INTEGER,
    usatf_valid BOOLEAN DEFAULT FALSE,
    usatf_age_verified BOOLEAN DEFAULT FALSE
)
"""


USATF_KEYS = [
    'Last Name',
    'First Name',
    'Date of Birth',
    'Sex',
    'Individual Membership Status',
    'Individual Membership Memb No.',
    'Date of Birth Verification Status'
]


def normalize(name):
    name = name.strip()
    return name.title() if name.islower() or name.isupper() else name


def normalize_usatf(row):
    norm_dict = {
        'First Name': normalize,
        'Last Name': normalize,
        'Date of Birth': date.fromisoformat,
        'Sex': normalize,
        'Individual Membership Status': lambda s: s.strip() == 'Current',
        'Individual Membership Memb No.': lambda n: int(n.strip()) if n else None,
        'Date of Birth Verification Status': lambda s: s.strip() == 'Current'
    }
    return [norm_dict[k](row[k]) for k in USATF_KEYS]


def adapt_date_iso(val):
    """Adapt datetime.date to ISO 8601 date."""
    return val.isoformat()


def adapt_datetime_epoch(val):
    """Adapt datetime.datetime to Unix timestamp."""
    return int(val.timestamp())


def convert_date(val):
    """Convert ISO 8601 date to datetime.date object."""
    return date.fromisoformat(val.decode())

def convert_timestamp(val):
    """Convert Unix epoch timestamp to datetime.datetime object."""
    return datetime.fromtimestamp(int(val))


sqlite3.register_adapter(date, adapt_date_iso)
sqlite3.register_adapter(datetime, adapt_datetime_epoch)
sqlite3.register_converter("date", convert_date)
sqlite3.register_converter("timestamp", convert_timestamp)


def prepare_db(db_filename):
    db = sqlite3.connect(
        db_filename,
        detect_types=sqlite3.PARSE_DECLTYPES|sqlite3.PARSE_COLNAMES
    )
    db.executescript(ROSTER_SCHEMA)
    return db


def find_dups(db: sqlite3.Connection):
    cur = db.execute(
        'SELECT ts_last, ts_first, COUNT(*) '
        'FROM roster GROUP BY ts_last, ts_first '
        'HAVING COUNT(*) > 1'
    )
    return cur.fetchall()


def init_roster(options):
    session_ts = datetime.now()
    print(f'Initializing/updating database from file: {options.ROSTER}')
    with closing(prepare_db(options.database)) as db:
        with open(options.ROSTER, 'rt') as roster_file:
            reader = csv.DictReader(roster_file)
            rows = [
                (
                    normalize(r['Last']),
                    normalize(r['First']),
                    date.fromisoformat(r['Birthdate']) if r['Birthdate'] else date(1900, 1, 1),
                    normalize(r['Gender']),
                    session_ts
                )
                for r in reader
            ]
            with db:
                db.executemany(
                    'INSERT OR IGNORE INTO roster(ts_last, ts_first, ts_dob, ts_gender, touched) '
                    'VALUES (?, ?, ?, ?, ?)',
                    rows
                )

            if (dups := find_dups(db)):
                print('WARNING: duplicate entries detected.  Use `rostmerg fixdups` to resolve.')
                if options.show_dups:
                    print('\tLast,First,Count')
                    for row in dups:
                        print('\t' + ','.join([str(f) for f in row]))


def fix_dups(options):
    with closing(prepare_db(options.database)) as db:
        dups = find_dups(db)
        if not dups:
            print('No duplicate entries found, exiting...')
            sys.exit()
        else:
            for (last, first, count) in dups:
                print(f'Found {count} duplicates for {last}, {first}:')
                rows = db.execute(
                    'SELECT * '
                    'FROM roster '
                    'WHERE ts_last = ? AND ts_first = ? '
                    'ORDER BY touched',
                    (last, first)
                ).fetchall()
                for i, row in enumerate(rows, start=1):
                    print(f'{i}) ' + ','.join([str(f) for f in row]))
                print(
                    '(R)etain duplicates; (M#) Merge, keeping TS birthdate from record #; '
                    '(D#) Drop record #; (Q)uit'
                )
                while True:
                    choice = input('? ').upper()
                    if choice and choice[0] in 'RMDQ':
                        if choice[0] == 'R':
                            break
                        elif choice[0] == 'M' and len(choice) > 1 and choice[1].isdigit():
                            i = int(choice[1]) - 1
                            if 0 <= i < len(rows):
                                with db:
                                    db.execute(
                                        'DELETE FROM roster '
                                        'WHERE ts_last = ? AND ts_first = ? AND ts_dob = ?',
                                        rows[i][:3]
                                    )
                                    db.execute(
                                        'UPDATE roster SET ts_dob = ? '
                                        'WHERE ts_last = ? AND ts_first = ? AND ts_dob = ?',
                                        [rows[i][2], *rows[0][:3]]
                                    )
                                break
                        elif choice[0] == 'D' and len(choice) > 1 and choice[1].isdigit():
                            i = int(choice[1]) - 1
                            if 0 <= i < len(rows):
                                with db:
                                    db.execute(
                                        'DELETE FROM roster '
                                        'WHERE ts_last = ? AND ts_first = ? AND ts_dob = ?',
                                        rows[i][:3]
                                    )
                                break
                        elif choice[0] == 'Q':
                            sys.exit()


def set_coaches(options):
    with closing(prepare_db(options.database)) as db:
        if options.clear:
            with db:
                db.execute('UPDATE roster SET is_coach = 0')
        else:
            rows = db.execute('SELECT ts_last, ts_first, ts_dob, is_coach FROM roster').fetchall()
            for (last, first, dob, coach) in rows:
                print(
                    f'Is ({last}, {first}, {str(dob)}) a coach? '
                    f'(current: {coach}) (Y)es, (N)o, (S)kip, (D)one'
                )
                while True:
                    choice = input('? ').upper()
                    if choice and choice[0] in 'YNSD':
                        if choice[0] == 'Y':
                            with db:
                                db.execute(
                                    'UPDATE roster SET is_coach = 1 '
                                    'WHERE ts_last = ? AND ts_first = ? AND ts_dob = ?',
                                    (last, first, dob)
                                )
                            break
                        elif choice[0] == 'N':
                            with db:
                                db.execute(
                                    'UPDATE roster SET is_coach = 0 '
                                    'WHERE ts_last = ? AND ts_first = ? AND ts_dob = ?',
                                    (last, first, dob)
                                )
                            break
                        elif choice[0] == 'S':
                            break
                        elif choice[0] == 'D':
                            return


def merge_usatf(options):
    with closing(prepare_db(options.database)) as db:
        with open(options.USATF, 'rt') as usatf_file:
            reader = csv.DictReader(usatf_file)
            data = [normalize_usatf(r) for r in reader]
        with db:
            db.executescript(USATF_SCHEMA)
        with db:
            db.executemany(
                'INSERT INTO usatf(usatf_last, usatf_first, usatf_dob, '
                'usatf_gender, usatf_valid, usatf_id, usatf_age_verified) '
                'VALUES (?, ?, ?, ?, ?, ?, ?)',
                data
            )
        team = db.execute(
            'SELECT ts_last, ts_first, ts_dob, ts_gender '
            'FROM roster '
            'WHERE is_coach = 0'
        ).fetchall()
        for runner in team:
            matches = db.execute(
                'SELECT * FROM usatf '
                'WHERE usatf_last = ? AND usatf_first = ? AND usatf_dob = ? AND usatf_gender = ?',
                runner
            ).fetchall()
            if len(matches) == 1:
                print(f'Found exact match for {runner}')
                with db:
                    params = [*matches[0][:3], *matches[0][4:], *runner[:3]]
                    db.execute(
                        'UPDATE roster '
                        'SET usatf_last = ?, usatf_first = ?, usatf_dob = ?, '
                        'usatf_id = ?, usatf_valid = ?, usatf_age_verified = ? '
                        'WHERE ts_last = ? AND ts_first = ? AND ts_dob = ?',
                        params
                    )
            elif len(matches) > 1:
                print(f'Found multiple matches for {runner}:')
                for i, match in enumerate(matches, start=1):
                    print(f'{i}) {match}')
            elif len(matches) == 0:
                print(f'Found zero matches for {runner}')
                print('Relaxing search...')
                matches = db.execute(
                    'SELECT * FROM usatf '
                    'WHERE (usatf_last = ? OR usatf_first = ? OR usatf_dob = ?) AND usatf_gender = ?',
                    runner
                ).fetchall()
                print('Partial matches:')
                for i, match in enumerate(matches, start=1):
                    print(f'\t{i}) {match}')
                print('Select matching record # or (S)kip')
                while True:
                    choice = input('? ')
                    if choice and choice[0].isdigit() and 0 <= (i := int(choice[0]) - 1) < len(matches):
                        with db:
                            params = [*matches[i][:3], *matches[i][4:], *runner[:3]]
                            db.execute(
                                'UPDATE roster '
                                'SET usatf_last = ?, usatf_first = ?, usatf_dob = ?, '
                                'usatf_id = ?, usatf_valid = ?, usatf_age_verified = ? '
                                'WHERE ts_last = ? AND ts_first = ? AND ts_dob = ?',
                                params
                            )
                        break
                    elif choice and choice[0].upper() == 'S':
                        break


def export_roster(options):
    if not options.year:
        options.year = datetime.now().year
    if not options.OUTPUT.endswith('.xlsx'):
        options.OUTPUT = options.OUTPUT + '.xlsx'
    HEADER = ('Last', 'First', 'USATF Age', 'USATF Status', 'USATF Num', 'Age Verified')
    GREEN_FILL = PatternFill(start_color='B7E1CD', end_color='B7E1CD', fill_type='solid')
    RED_FILL = PatternFill(start_color='E06666', end_color='E06666', fill_type='solid')
    with closing(prepare_db(options.database)) as db:
        rows = db.execute(
            'SELECT ts_last, ts_first, ts_dob, usatf_valid, usatf_id, usatf_age_verified '
            'FROM roster '
            'WHERE is_coach = 0 '
            'ORDER BY ts_last, ts_first'
        ).fetchall()

        wb = Workbook()
        wb.active.title = 'registration'
        ws = wb.active
        for c, h in zip('ABCDEF', HEADER):
            ws[f'{c}1'] = CellRichText(TextBlock(InlineFont(b=True), h))
            ws.column_dimensions[c].width = 14.29
        ws.freeze_panes = 'A2'
        for row in rows:
            r = list(row[:2])
            r.append(options.year - row[2].year)
            r.append('Current' if row[3] else 'Not Assoc')
            r.append(row[4])
            r.append('Current' if row[5] else '')
            ws.append(r)
        for col in (f'D2:D{len(rows)+1}', f'F2:F{len(rows)+1}'):
            ws.conditional_formatting.add(
                col,
                Rule(
                    type='containsText', operator='containsText', text='Current',
                    formula=[f'NOT(ISERROR(SEARCH("Current",{col[0]}2)))'],
                    stopIfTrue=False,
                    dxf=DifferentialStyle(fill=GREEN_FILL)
                )
            )
            ws.conditional_formatting.add(
                col,
                Rule(
                    type='notContainsText', operator='notContains', text='Current',
                    formula=[f'ISERROR(SEARCH("Current",{col[0]}2))'],
                    stopIfTrue=False,
                    dxf=DifferentialStyle(fill=RED_FILL)
                )
            )
        wb.save(options.OUTPUT)
        wb.close()


if __name__ == '__main__':
    import argparse

    parser = argparse.ArgumentParser(description="A program to merge TeamSnap and USATF roster information")
    parser.add_argument(
        '-d', '--database', type=str, default='roster.db',
        help='Roster database (defaults to "roster.db")'
    )
    subparsers = parser.add_subparsers(required=True)
    init = subparsers.add_parser('init', help='Initialize (or update a roster)')
    init.add_argument('ROSTER', type=str, help='A CSV file containing a roster exported from TeamSnap')
    init.add_argument('-s', '--show-dups', action='store_true', help='Show which names have duplicate entries')
    init.set_defaults(func=init_roster)

    fix = subparsers.add_parser('fixdups', help='Resolve duplicate entries in the database')
    fix.set_defaults(func=fix_dups)

    coach = subparsers.add_parser('coaches', help='Set (or clear) coaching entries')
    coach.add_argument('-c', '--clear', action='store_true', help='Clear all coaching flags')
    coach.set_defaults(func=set_coaches)

    merge = subparsers.add_parser('merge', help='Merge USATF membership data')
    merge.add_argument('USATF', type=str, help='A CSV file containing USATF membership and age verification data')
    merge.set_defaults(func=merge_usatf)

    export = subparsers.add_parser('export', help='Export the roster')
    export.add_argument('OUTPUT', type=str, help='Name of the exported roster file.')
    export.add_argument(
        '-y', '--year', type=int, default=None,
        help='Year to be used for calculating the age, defaults to the current year'
    )
    export.set_defaults(func=export_roster)

    options = parser.parse_args()
    options.func(options)
